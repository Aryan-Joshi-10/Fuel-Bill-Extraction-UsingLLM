import os
import json
import time
import traceback
import logging
from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv
from PIL import Image
import openpyxl
import pdfplumber
import google.generativeai as genai
from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from typing import List, Optional
from datetime import datetime
from io import BytesIO
from pydantic import BaseModel
import pathlib
from werkzeug.utils import secure_filename

# === Initial Setup ===
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
os.makedirs('logs', exist_ok=True)
file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240, backupCount=10)
file_handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'))
logger.addHandler(file_handler)

# Configure Gemini
try:
    genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))
    model = genai.GenerativeModel("gemini-2.5-flash")
except Exception as e:
    logger.error(f"Failed to configure Gemini: {str(e)}")
    raise

# FastAPI app initialization
app = FastAPI(title="Income Tax Filing Automation API", version="1.0.0")

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure directories
UPLOAD_FOLDER = os.getenv("UPLOAD_FOLDER", "uploads")
DATA_DIR = "data"
EXCEL_OUTPUT_DIR = "outputs"
MAX_CONTENT_LENGTH = int(os.getenv("MAX_CONTENT_LENGTH", 50 * 1024 * 1024))
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)

# File paths for storing data
DECLARATION_FILE = os.path.join(DATA_DIR, "tax_declaration.json")
DRIVER_SALARY_FILE = os.path.join(DATA_DIR, "driver_salary.json")

# === Pydantic Models ===
class TaxDeclaration(BaseModel):
    financial_year: str
    pan_number: str
    name: str
    declared_fuel_amount: float
    declared_driver_salary: float
    other_declarations: Optional[dict] = {}
    notes: Optional[str] = ""

class DriverSalary(BaseModel):
    driver_name: str
    vehicle_number: Optional[str] = None
    driver_license_number: Optional[str] = None
    monthly_salary: float
    months_worked: int
    total_salary: Optional[float] = None
    salary_slips: Optional[List[str]] = []
    notes: Optional[str] = ""

class TaxCalculationRequest(BaseModel):
    tax_regime: str  # 'old' or 'new'
    financial_year: str
    income: dict
    deductions: dict

# === Helper Functions ===
def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def cleanup_old_files():
    """Removes uploaded files older than 24 hours"""
    try:
        current_time = time.time()
        for filename in os.listdir(UPLOAD_FOLDER):
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.getmtime(filepath) < current_time - 86400:
                os.remove(filepath)
                logger.info(f"Cleaned up old file: {filename}")
    except Exception as e:
        logger.error(f"Error during cleanup: {str(e)}")

def load_json_file(filepath: str, default: dict = None):
    """Load JSON data from file"""
    if default is None:
        default = {}
    try:
        if os.path.exists(filepath):
            with open(filepath, 'r') as f:
                return json.load(f)
        return default
    except Exception as e:
        logger.error(f"Error loading {filepath}: {str(e)}")
        return default

def save_json_file(filepath: str, data: dict):
    """Save JSON data to file"""
    try:
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2)
        return True
    except Exception as e:
        logger.error(f"Error saving {filepath}: {str(e)}")
        return False

# === Prompt for Gemini ===
prompt = """
You are a vision-language model tasked with extracting structured information from petrol or diesel bills. These bills may be in English, Hindi, or Marathi.
Your goal is to extract and return the following details in JSON format:

- **Petrol Pump Name**: The topmost prominent text, usually representing the petrol pump or brand name (e.g., "Tungar Petroleum").
- **Date**: The date of the transaction. It may appear near the bill number or be labeled as "Date:", "दिनांक", or "दि.". Return the date strictly in DD/MM/YYYY format.
- **Product**: Identify the type of fuel sold. Extract **only** the word "Petrol" or "Diesel". Do not include any numbers, prices, or quantities. Choose strictly between:
  - "Petrol"
  - "Diesel"
- **Volume(L)**: The value mentioned next to the label "VOLUME" or its equivalent.
- **Rate per Litre**: The rate of the fuel per litre. This is usually in the third column of a price table. For example, if shown as "91\n74", convert it to "91.74".
- **Total Amount (Rs)**: The final amount payable, generally found near the bottom-right under the label "AMOUNT" or "Rs." or "LKR". - When the Total Amount is not explicitly found, try to estimate based on tabular layout. For example, the last value in the third column of the price table usually corresponds to the final amount.

**Additional Instructions:**
- Translate all extracted information into English.
- If a particular field is missing or unclear in the image, leave its value as an empty string.
- Return the final result strictly in the following JSON format:

```json
{
  "Petrol Pump Name": "",
  "Date": "",
  "Product": "",
  "Volume(L)": "",
  "Rate per Litre": "",
  "Total Amount (Rs)": ""
}
```
"""

# === Core Processing ===
def process_image(image, fuel_bill_no):
    try:
        response = model.generate_content([prompt, image])
        content = response.text.strip()

        if content.startswith("```json"):
            content = content[7:-3].strip()
        elif content.startswith("```"):
            content = content[3:-3].strip()

        data = json.loads(content)

        # Estimate total if missing
        if not data.get("Total Amount (Rs)", "").strip() and data.get("Volume(L)") and data.get("Rate per Litre"):
            try:
                total = round(float(data["Volume(L)"]) * float(data["Rate per Litre"]), 2)
                data["Total Amount (Rs)"] = str(total)
            except ValueError:
                pass

        return {"file": fuel_bill_no, "data": data}

    except Exception as e:
        return {"file": fuel_bill_no, "error": str(e)}

# === Routes ===

# Health Check
@app.get("/health")
async def health_check():
    try:
        test_response = model.generate_content("Test")
        gemini_status = "healthy" if test_response else "unhealthy"
        upload_dir_status = "healthy" if os.access(UPLOAD_FOLDER, os.W_OK) else "unhealthy"

        return JSONResponse(
            status_code=200,
            content={
                "status": "healthy",
                "components": {
                    "gemini_api": gemini_status,
                    "upload_directory": upload_dir_status,
                    "pdf_processing": "healthy"
                },
            },
        )
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        return JSONResponse(status_code=500, content={"status": "unhealthy", "error": str(e)})

# === Income Tax Declaration Routes ===
@app.get("/api/declaration")
async def get_declaration():
    """Get income tax declaration details"""
    data = load_json_file(DECLARATION_FILE)
    return JSONResponse(status_code=200, content={"success": True, "data": data})

@app.post("/api/declaration")
async def save_declaration(declaration: TaxDeclaration):
    """Save income tax declaration details"""
    data = declaration.dict()
    data["last_updated"] = datetime.now().isoformat()
    
    if save_json_file(DECLARATION_FILE, data):
        return JSONResponse(status_code=200, content={"success": True, "message": "Declaration saved successfully", "data": data})
    else:
        raise HTTPException(status_code=500, detail="Failed to save declaration")

@app.put("/api/declaration")
async def update_declaration(declaration: TaxDeclaration):
    """Update income tax declaration details"""
    data = declaration.dict()
    data["last_updated"] = datetime.now().isoformat()
    
    if save_json_file(DECLARATION_FILE, data):
        return JSONResponse(status_code=200, content={"success": True, "message": "Declaration updated successfully", "data": data})
    else:
        raise HTTPException(status_code=500, detail="Failed to update declaration")

# === Fuel Bills Routes ===
@app.post("/api/fuel-bills/upload")
async def upload_fuel_bills(files: List[UploadFile] = File(...)):
    """Upload and extract data from fuel bills"""
    try:
        cleanup_old_files()

        if not files:
            raise HTTPException(status_code=400, detail="No files uploaded")

        results = []
        all_bills_data = []

        for file in files:
            filename = file.filename
            if not allowed_file(filename):
                results.append({
                    "file": filename,
                    "error": "Invalid file type. Allowed: PDF, PNG, JPG, JPEG"
                })
                continue

            # Use secure filename and ensure unique name
            safe_filename = secure_filename(filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            filepath = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{safe_filename}")
            
            # Save file temporarily
            try:
                content = await file.read()
                if len(content) == 0:
                    results.append({
                        "file": filename,
                        "error": "File is empty"
                    })
                    continue
                    
                with open(filepath, "wb") as buffer:
                    buffer.write(content)
                logger.info(f"Saved file: {filepath} ({len(content)} bytes)")
            except Exception as e:
                logger.error(f"Error saving file {filename}: {str(e)}")
                results.append({
                    "file": filename,
                    "error": f"Error saving file: {str(e)}"
                })
                continue

            images_to_process = []

            try:
                if filename.lower().endswith(".pdf"):
                    with pdfplumber.open(filepath) as pdf:
                        if not pdf.pages:
                            raise Exception("No pages found in PDF")
                        for page in pdf.pages:
                            img = page.to_image(resolution=300).original
                            images_to_process.append(img)
                else:
                    img = Image.open(filepath)
                    images_to_process.append(img)
            except Exception as e:
                logger.error(f"Error processing {filename}: {str(e)}")
                results.append({"file": filename, "error": str(e)})
                if os.path.exists(filepath):
                    os.remove(filepath)
                continue

            for i, img in enumerate(images_to_process):
                fuel_bill_no = f"{os.path.splitext(filename)[0]}_page{i + 1}" if len(images_to_process) > 1 else os.path.splitext(filename)[0]
                result = process_image(img, fuel_bill_no)
                results.append(result)
                
                if "data" in result:
                    all_bills_data.append(result["data"])

            # Clean up file
            if os.path.exists(filepath):
                os.remove(filepath)

        # Calculate total fuel cost
        total_fuel_cost = 0
        for bill in all_bills_data:
            try:
                amount = float(bill.get("Total Amount (Rs)", "0") or "0")
                total_fuel_cost += amount
            except (ValueError, TypeError):
                pass

        print(f"Testing:  {results}");
        return JSONResponse(
            status_code=200,
            content={
                "success": True,
                "results": results,
                "total_bills": len(all_bills_data),
                "total_fuel_cost": round(total_fuel_cost, 2),
                "bills_data": all_bills_data
            }
        )

    except Exception as e:
        logger.error(f"Error in upload: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e), "traceback": traceback.format_exc()},
        )
# @app.post("/api/fuel-bills/upload")
# async def upload_fuel_bills(files: List[UploadFile] = File(...)):
#     all_bills_data = []
#     print("Testing the upload route")
#     for file in files:
#         contents = await file.read()
#         print(f"✅ Received {file.filename}, size = {len(contents)} bytes")

#     return JSONResponse(
#         content={
#             "success": True,
#             "total_files": len(files)
#         }
#     )

@app.post("/api/fuel-bills/generate-excel")
async def generate_fuel_bills_excel(bills_data: List[dict]):
    """Generate Excel file from fuel bills data"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fuel Bills"
        
        # Headers
        headers = ["Fuel_bill_No.", "Petrol Pump Name", "Date", "Product", "Volume(L)", "Rate per Litre", "Total Amount (Rs)"]
        ws.append(headers)
        
        # Style headers
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data
        total_amount = 0
        for idx, bill in enumerate(bills_data):
            # Handle both direct dict and nested data structure
            bill_data = bill.get("data", bill) if "data" in bill else bill
            bill_no = bill.get("file", f"Bill_{idx+1}")
            row = [
                bill_no,
                bill_data.get("Petrol Pump Name", ""),
                bill_data.get("Date", ""),
                bill_data.get("Product", ""),
                bill_data.get("Volume(L)", ""),
                bill_data.get("Rate per Litre", ""),
                bill_data.get("Total Amount (Rs)", "")
            ]
            ws.append(row)
            
            try:
                amount_str = bill_data.get("Total Amount (Rs)", "0") or "0"
                amount = float(amount_str)
                total_amount += amount
            except (ValueError, TypeError):
                pass
        
        # Add total row
        ws.append([])
        ws.append(["TOTAL", "", "", "", "", "", f"₹{round(total_amount, 2)}"])
        total_cell = ws.cell(row=ws.max_row, column=7)
        total_cell.font = Font(bold=True, size=12)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"fuel_bills_{timestamp}.xlsx"
        filepath = os.path.join(EXCEL_OUTPUT_DIR, filename)
        
        # Save to disk
        wb.save(filepath)
        
        return FileResponse(
            path=filepath,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")

# === Driver Salary Routes ===
@app.get("/api/driver-salary")
async def get_driver_salary():
    """Get driver salary details"""
    data = load_json_file(DRIVER_SALARY_FILE)
    return JSONResponse(status_code=200, content={"success": True, "data": data})

@app.post("/api/driver-salary")
async def save_driver_salary(salary: DriverSalary):
    """Save driver salary details"""
    data = salary.dict()
    
    # Calculate total if not provided
    if data.get("total_salary") is None:
        data["total_salary"] = data["monthly_salary"] * data["months_worked"]
    
    data["last_updated"] = datetime.now().isoformat()
    
    if save_json_file(DRIVER_SALARY_FILE, data):
        return JSONResponse(status_code=200, content={"success": True, "message": "Driver salary saved successfully", "data": data})
    else:
        raise HTTPException(status_code=500, detail="Failed to save driver salary")

@app.put("/api/driver-salary")
async def update_driver_salary(salary: DriverSalary):
    """Update driver salary details"""
    data = salary.dict()
    
    # Calculate total if not provided
    if data.get("total_salary") is None:
        data["total_salary"] = data["monthly_salary"] * data["months_worked"]
    
    data["last_updated"] = datetime.now().isoformat()
    
    if save_json_file(DRIVER_SALARY_FILE, data):
        return JSONResponse(status_code=200, content={"success": True, "message": "Driver salary updated successfully", "data": data})
    else:
        raise HTTPException(status_code=500, detail="Failed to update driver salary")

@app.post("/api/driver-salary/validate")
async def validate_driver_salary():
    """Validate driver salary against declared amount"""
    try:
        declaration = load_json_file(DECLARATION_FILE)
        salary_data = load_json_file(DRIVER_SALARY_FILE)
        
        # Check if files exist and have data
        if not declaration or declaration == {}:
            raise HTTPException(
                status_code=400, 
                detail="Tax declaration not found. Please fill in the Tax Declaration form first."
            )
        
        if not salary_data or salary_data == {}:
            raise HTTPException(
                status_code=400, 
                detail="Driver salary data not found. Please save driver salary details first."
            )
        
        declared_amount = float(declaration.get("declared_driver_salary", 0) or 0)
        calculated_amount = float(salary_data.get("total_salary", 0) or 0)
        
        if calculated_amount == 0:
            raise HTTPException(
                status_code=400,
                detail="Total salary is zero. Please ensure monthly salary and months worked are set."
            )
        
        difference = calculated_amount - declared_amount
        is_valid = abs(difference) <= 100  # Allow 100 Rs difference
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True,
                "declared_amount": declared_amount,
                "calculated_amount": calculated_amount,
                "difference": round(difference, 2),
                "is_valid": is_valid,
                "message": "Valid" if is_valid else f"Difference of ₹{abs(difference):.2f} found"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error validating driver salary: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error validating driver salary: {str(e)}")

# === Validation Route ===
@app.get("/api/validate")
async def validate_all():
    """Validate all declared amounts against calculated values"""
    declaration = load_json_file(DECLARATION_FILE)
    salary_data = load_json_file(DRIVER_SALARY_FILE)
    
    if not declaration:
        raise HTTPException(status_code=400, detail="No declaration found")
    
    results = {
        "declaration": declaration,
        "validations": {}
    }
    
    # Validate driver salary
    if salary_data:
        declared_salary = declaration.get("declared_driver_salary", 0)
        calculated_salary = salary_data.get("total_salary", 0)
        salary_diff = calculated_salary - declared_salary
        
        results["validations"]["driver_salary"] = {
            "declared": declared_salary,
            "calculated": calculated_salary,
            "difference": round(salary_diff, 2),
            "is_valid": abs(salary_diff) <= 100
        }
    
    # Validate fuel bills
    # Try multiple methods to get fuel bills total
    total_fuel_cost = 0
    fuel_bills_count = 0
    
    # Option 1: Try serverless API endpoint if available
    serverless_api_url = os.getenv('SERVERLESS_API_URL')
    if serverless_api_url:
        try:
            import httpx
            with httpx.Client(timeout=10.0) as client:
                resp = client.get(f"{serverless_api_url}/fuel-bills-total")
                if resp.status_code == 200:
                    api_data = resp.json()
                    total_fuel_cost = float(api_data.get('total', 0) or 0)
                    fuel_bills_count = int(api_data.get('count', 0) or 0)
                    logger.info(f"Got fuel bills total from serverless API: {total_fuel_cost} from {fuel_bills_count} bills")
        except ImportError:
            logger.info("httpx not installed, skipping serverless API call")
        except Exception as e:
            logger.warning(f"Serverless API not available: {str(e)}")
    
    # Option 2: Try DynamoDB directly (if serverless API didn't work)
    if total_fuel_cost == 0:
        try:
            import boto3
            dynamodb = boto3.client('dynamodb', region_name=os.getenv('AWS_REGION', 'ap-south-1'))
            table_name = os.getenv('DDB_TABLE', 'FuelBillJobs')
            
            logger.info(f"Scanning DynamoDB table: {table_name}")
            # Scan DynamoDB table for all completed fuel bills
            response = dynamodb.scan(
                TableName=table_name,
                FilterExpression='job_status = :status',
                ExpressionAttributeValues={':status': {'S': 'COMPLETED'}}
            )
            
            logger.info(f"Found {len(response.get('Items', []))} completed items in DynamoDB")
            
            for item in response.get('Items', []):
                if 'data' in item:
                    try:
                        data_str = item['data']['S']
                        data = json.loads(data_str)
                        fuel_bills_count += 1
                        
                        # Handle both raw_text (Gemini response) and parsed data
                        if isinstance(data, dict):
                            amount = 0
                            if 'raw_text' in data:
                                # Parse JSON from raw_text if it's in markdown code block
                                raw = data['raw_text']
                                if '```json' in raw or '```' in raw:
                                    json_start = raw.find('{')
                                    json_end = raw.rfind('}') + 1
                                    if json_start >= 0 and json_end > json_start:
                                        try:
                                            parsed = json.loads(raw[json_start:json_end])
                                            amount = float(parsed.get('total_amount', 0) or parsed.get('Total Amount (Rs)', 0) or 0)
                                        except:
                                            pass
                                elif '{' in raw:
                                    json_start = raw.find('{')
                                    json_end = raw.rfind('}') + 1
                                    if json_start >= 0 and json_end > json_start:
                                        try:
                                            parsed = json.loads(raw[json_start:json_end])
                                            amount = float(parsed.get('total_amount', 0) or parsed.get('Total Amount (Rs)', 0) or 0)
                                        except:
                                            pass
                            else:
                                # Direct data structure - check multiple possible field names
                                amount = float(
                                    data.get('total_amount', 0) or 
                                    data.get('Total Amount (Rs)', 0) or 
                                    data.get('total', 0) or 
                                    0
                                )
                            
                            total_fuel_cost += amount
                            logger.debug(f"Added amount: {amount}, running total: {total_fuel_cost}")
                    except (json.JSONDecodeError, ValueError, KeyError, TypeError) as e:
                        logger.warning(f"Error parsing fuel bill data: {str(e)}, item: {item.get('file_key', {}).get('S', 'unknown')}")
                        continue
            
            logger.info(f"Total fuel cost calculated from DynamoDB: {total_fuel_cost} from {fuel_bills_count} bills")
        except ImportError:
            logger.warning("boto3 not installed, cannot access DynamoDB directly")
        except Exception as e:
            logger.error(f"DynamoDB error: {str(e)}", exc_info=True)
    
    # Only add fuel bills validation if we have calculated total
    if total_fuel_cost > 0:
        declared_fuel = declaration.get("declared_fuel_amount", 0) or 0
        fuel_diff = total_fuel_cost - declared_fuel
        
        results["validations"]["fuel_bills"] = {
            "declared": declared_fuel,
            "calculated": round(total_fuel_cost, 2),
            "difference": round(fuel_diff, 2),
            "is_valid": abs(fuel_diff) <= 100,
            "bills_count": fuel_bills_count
        }
    else:
        logger.warning("No fuel bills data found for validation")
    
    return JSONResponse(status_code=200, content={"success": True, "data": results})

# === Tax Calculator Routes ===
def calculate_tax_old_regime(taxable_income: float, financial_year: str = "2024-25"):
    """Calculate tax for old regime (FY 2024-25)"""
    # Tax slabs for FY 2024-25 (AY 2025-26)
    slabs = [
        {"min": 0, "max": 250000, "rate": 0},
        {"min": 250000, "max": 500000, "rate": 0.05},
        {"min": 500000, "max": 1000000, "rate": 0.20},
        {"min": 1000000, "max": float('inf'), "rate": 0.30}
    ]
    
    tax = 0
    tax_breakdown = []
    remaining_income = taxable_income
    
    for slab in slabs:
        if remaining_income <= 0:
            break
        
        slab_income = min(remaining_income, slab["max"] - slab["min"])
        if slab_income > 0:
            slab_tax = slab_income * slab["rate"]
            tax += slab_tax
            tax_breakdown.append({
                "range": f"₹{slab['min']:,.0f} - ₹{min(slab['max'], taxable_income):,.0f}",
                "tax": round(slab_tax, 2)
            })
            remaining_income -= slab_income
    
    return tax, tax_breakdown

def calculate_tax_new_regime(taxable_income: float, financial_year: str = "2024-25"):
    """Calculate tax for new regime (FY 2024-25)"""
    # Tax slabs for new regime FY 2024-25
    slabs = [
        {"min": 0, "max": 300000, "rate": 0},
        {"min": 300000, "max": 700000, "rate": 0.05},
        {"min": 700000, "max": 1000000, "rate": 0.10},
        {"min": 1000000, "max": 1200000, "rate": 0.15},
        {"min": 1200000, "max": 1500000, "rate": 0.20},
        {"min": 1500000, "max": float('inf'), "rate": 0.30}
    ]
    
    tax = 0
    tax_breakdown = []
    remaining_income = taxable_income
    
    for slab in slabs:
        if remaining_income <= 0:
            break
        
        slab_income = min(remaining_income, slab["max"] - slab["min"])
        if slab_income > 0:
            slab_tax = slab_income * slab["rate"]
            tax += slab_tax
            tax_breakdown.append({
                "range": f"₹{slab['min']:,.0f} - ₹{min(slab['max'], taxable_income):,.0f}",
                "tax": round(slab_tax, 2)
            })
            remaining_income -= slab_income
    
    return tax, tax_breakdown

def get_tax_savings_suggestions(income: float, deductions: dict, tax_regime: str):
    """Generate tax savings suggestions"""
    suggestions = []
    
    if tax_regime == 'old':
        # Check 80C utilization
        section_80c = deductions.get('section_80c', 0)
        if section_80c < 150000 and income > 500000:
            remaining_80c = 150000 - section_80c
            suggestions.append(
                f"Invest ₹{remaining_80c:,.0f} more in Section 80C (ELSS, PPF, NSC) to save up to ₹{remaining_80c * 0.30:,.0f} in tax"
            )
        
        # Check 80D utilization
        section_80d = deductions.get('section_80d', 0)
        if section_80d < 100000 and income > 500000:
            remaining_80d = 100000 - section_80d
            suggestions.append(
                f"Consider health insurance premium of ₹{remaining_80d:,.0f} under Section 80D to save up to ₹{remaining_80d * 0.30:,.0f} in tax"
            )
        
        # Check home loan interest
        home_loan = deductions.get('home_loan_interest_24b', 0)
        if home_loan < 200000 and income > 1000000:
            remaining_home = 200000 - home_loan
            suggestions.append(
                f"Home loan interest deduction available up to ₹{remaining_home:,.0f} under Section 24(b)"
            )
    
    # General suggestions
    if income > 1000000:
        suggestions.append("Consider consulting a tax advisor for advanced tax planning strategies")
    
    if deductions.get('section_80g', 0) == 0 and income > 500000:
        suggestions.append("Donations to eligible charities under Section 80G can provide tax benefits")
    
    return suggestions

@app.post("/api/tax-calculator/calculate")
async def calculate_tax(request: TaxCalculationRequest):
    """Calculate tax liability based on income and deductions"""
    try:
        income_data = request.income
        deductions_data = request.deductions
        
        # Calculate gross total income
        gross_total_income = (
            income_data.get('salary', 0) +
            income_data.get('business', 0) +
            income_data.get('capital_gains', 0) +
            income_data.get('other', 0)
        )
        
        if gross_total_income <= 0:
            raise HTTPException(status_code=400, detail="Total income must be greater than zero")
        
        # Calculate total deductions
        total_deductions = 0
        if request.tax_regime == 'old':
            # Apply deduction limits
            section_80c = min(deductions_data.get('section_80c', 0), 150000)
            section_80d = min(deductions_data.get('section_80d', 0), 100000)
            home_loan_24b = min(deductions_data.get('home_loan_interest_24b', 0), 200000)
            section_80tta = min(deductions_data.get('section_80tta', 0), 10000)
            
            total_deductions = (
                section_80c +
                section_80d +
                deductions_data.get('hra', 0) +
                home_loan_24b +
                deductions_data.get('section_80g', 0) +
                deductions_data.get('section_80e', 0) +
                section_80tta +
                deductions_data.get('standard_deduction', 50000) +
                deductions_data.get('other_deductions', 0)
            )
        else:
            # New regime: only standard deduction
            total_deductions = deductions_data.get('standard_deduction', 50000)
        
        # Calculate taxable income
        taxable_income = max(0, gross_total_income - total_deductions)
        
        # Calculate tax based on regime
        if request.tax_regime == 'old':
            total_tax, tax_breakdown = calculate_tax_old_regime(taxable_income, request.financial_year)
        else:
            total_tax, tax_breakdown = calculate_tax_new_regime(taxable_income, request.financial_year)
        
        # Calculate cess (4% of tax)
        cess = total_tax * 0.04
        
        # Get tax savings suggestions
        suggestions = get_tax_savings_suggestions(gross_total_income, deductions_data, request.tax_regime)
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True,
                "gross_total_income": round(gross_total_income, 2),
                "total_deductions": round(total_deductions, 2),
                "taxable_income": round(taxable_income, 2),
                "total_tax": round(total_tax, 2),
                "cess": round(cess, 2),
                "tax_breakdown": tax_breakdown,
                "tax_savings_suggestions": suggestions,
                "tax_regime": request.tax_regime,
                "financial_year": request.financial_year
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error calculating tax: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error calculating tax: {str(e)}")

@app.post("/api/tax-calculator/compare-declaration")
async def compare_tax_with_declaration(request: TaxCalculationRequest):
    """Calculate tax and compare with declared amounts from tax declaration"""
    try:
        # First calculate the tax
        income_data = request.income
        deductions_data = request.deductions
        
        gross_total_income = (
            income_data.get('salary', 0) +
            income_data.get('business', 0) +
            income_data.get('capital_gains', 0) +
            income_data.get('other', 0)
        )
        
        if gross_total_income <= 0:
            raise HTTPException(status_code=400, detail="Total income must be greater than zero")
        
        # Load declaration data
        declaration = load_json_file(DECLARATION_FILE)
        if not declaration or declaration == {}:
            raise HTTPException(
                status_code=400,
                detail="Tax declaration not found. Please fill in the Tax Declaration form first."
            )
        
        # Calculate total deductions
        total_deductions = 0
        if request.tax_regime == 'old':
            section_80c = min(deductions_data.get('section_80c', 0), 150000)
            section_80d = min(deductions_data.get('section_80d', 0), 100000)
            home_loan_24b = min(deductions_data.get('home_loan_interest_24b', 0), 200000)
            section_80tta = min(deductions_data.get('section_80tta', 0), 10000)
            
            total_deductions = (
                section_80c +
                section_80d +
                deductions_data.get('hra', 0) +
                home_loan_24b +
                deductions_data.get('section_80g', 0) +
                deductions_data.get('section_80e', 0) +
                section_80tta +
                deductions_data.get('standard_deduction', 50000) +
                deductions_data.get('other_deductions', 0)
            )
        else:
            total_deductions = deductions_data.get('standard_deduction', 50000)
        
        taxable_income = max(0, gross_total_income - total_deductions)
        
        # Calculate tax
        if request.tax_regime == 'old':
            total_tax, tax_breakdown = calculate_tax_old_regime(taxable_income, request.financial_year)
        else:
            total_tax, tax_breakdown = calculate_tax_new_regime(taxable_income, request.financial_year)
        
        cess = total_tax * 0.04
        
        # Get declared amounts (approximate from fuel and driver salary)
        declared_fuel = float(declaration.get("declared_fuel_amount", 0) or 0)
        declared_driver_salary = float(declaration.get("declared_driver_salary", 0) or 0)
        
        # Estimate declared total income (this is approximate)
        # In real scenario, you'd have a separate field for total income in declaration
        declared_total_income = declared_fuel + declared_driver_salary + income_data.get('salary', 0)
        
        # Comparison
        comparison = {
            "declared_total_income": declared_total_income,
            "calculated_total_income": gross_total_income,
            "difference": gross_total_income - declared_total_income
        }
        
        suggestions = get_tax_savings_suggestions(gross_total_income, deductions_data, request.tax_regime)
        
        return JSONResponse(
            status_code=200,
            content={
                "success": True,
                "gross_total_income": round(gross_total_income, 2),
                "total_deductions": round(total_deductions, 2),
                "taxable_income": round(taxable_income, 2),
                "total_tax": round(total_tax, 2),
                "cess": round(cess, 2),
                "tax_breakdown": tax_breakdown,
                "tax_savings_suggestions": suggestions,
                "comparison": comparison,
                "tax_regime": request.tax_regime,
                "financial_year": request.financial_year
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error comparing tax with declaration: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error comparing tax: {str(e)}")

# Serve frontend static files (optional - for serving from same server)
# FRONTEND_DIR = pathlib.Path(__file__).parent.parent / "frontend"
# if FRONTEND_DIR.exists():
#     app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")

# @app.get("/", response_class=HTMLResponse)
# async def serve_frontend():
#     """Serve the frontend HTML file"""
#     frontend_path = FRONTEND_DIR / "index.html"
#     if frontend_path.exists():
#         with open(frontend_path, "r") as f:
#             return HTMLResponse(content=f.read())
#     return HTMLResponse(content="<h1>Frontend not found. Please open index.html directly.</h1>")

FRONTEND_DIR = pathlib.Path(__file__).parent.parent / "frontend"

if not FRONTEND_DIR.exists():
    logger.error(f"Frontend directory not found. Create it at: {FRONTEND_DIR}")
else:
    logger.info(f"Serving frontend from: {FRONTEND_DIR}")
    # app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")
    app.mount("/app", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")


# === Startup Event ===
@app.on_event("startup")
def startup_event():
    logger.info("🚀 FastAPI server started successfully.")
    logger.info(f"📁 Data directory: {DATA_DIR}")
    logger.info(f"📁 Upload directory: {UPLOAD_FOLDER}")
    logger.info(f"📁 Output directory: {EXCEL_OUTPUT_DIR}")
    logger.info(f"🌐 API available at: http://localhost:8000")
    logger.info(f"📄 API docs available at: http://localhost:8000/docs")

# === Main (for local run) ===
if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    logger.info(f"Starting FastAPI on port {port}")
    # Avoid reloads when writing to runtime/output directories during uploads
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=port,
        reload=True,
        reload_excludes=["uploads/*", "outputs/*", "logs/*"]
    )

